import openpyxl
from io import BytesIO
import datetime


def export_order(order_data: list[dict], include_mapping: bool = True) -> bytes:
    """Export processed order to Excel for 1C using openpyxl directly"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказ"

    # Headers
    headers = ["Артикул клиента", "Название клиента", "Количество"]
    if include_mapping:
        headers.extend(
            [
                "Артикул поставщика",
                "Название поставщика",
                "Упаковка",
                "Совпадение %",
                "Тип маппинга",
                "Требует проверки",
            ]
        )

    ws.append(headers)

    # Apply header style
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Data
    for item in order_data:
        qty = item.get("quantity", 1)
        row = [item.get("client_sku", ""), item.get("client_name", ""), qty]

        if include_mapping:
            match = item.get("match", {})
            row.extend(
                [
                    match.get("product_sku", ""),
                    match.get("product_name", ""),
                    match.get("pack_qty", 1),
                    match.get("confidence", 0),
                    match.get("match_type", ""),
                    "Да" if match.get("needs_review", True) else "Нет",
                ]
            )

        ws.append(row)

    # Adjust column widths
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value)))
                )

    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def test_export():
    # Mock data
    data = [
        {
            "client_sku": "123",
            "client_name": "Test Item",
            "quantity": 5,
            "match": {
                "product_sku": "JAKKO-123",
                "product_name": "Jakko Item",
                "pack_qty": 10,
                "confidence": 95,
                "match_type": "exact",
                "needs_review": False,
            },
        },
        {
            "client_sku": "456",
            "client_name": "Unknown Item",
            "quantity": 1,
            "match": {
                "product_sku": "",
                "product_name": "",
                "pack_qty": 1,
                "confidence": 0,
                "match_type": "not_found",
                "needs_review": True,
            },
        },
    ]

    print("Generating Excel...")
    excel_bytes = export_order(data)
    print(f"Generated {len(excel_bytes)} bytes")

    # Verify we can read it back
    print("Verifying readback...")
    wb = openpyxl.load_workbook(BytesIO(excel_bytes), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    print(f"Read {len(rows)} rows")
    assert len(rows) == 3  # Header + 2 data rows
    assert rows[1][0] == "123"
    assert rows[1][3] == "JAKKO-123"

    print("✅ Export verification passed!")


if __name__ == "__main__":
    test_export()
