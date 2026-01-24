import openpyxl
from io import BytesIO
import datetime


def create_sample_excel():
    """Create a sample Excel file in memory for testing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Order"
    # Headers
    ws.append(["Артикул", "Название", "Количество", "Цена"])
    # Data
    ws.append(["101105032R", "Муфта ППР 32", 10, 50.5])
    ws.append(["202132110K", "Ревизия кан. 110", 5, 200.0])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def parse_excel_file(file, filename: str):
    """
    Parse Excel file using openpyxl (replacing pandas.read_excel).
    Returns list of dicts: {'sku': ..., 'name': ..., 'qty': ...}
    """
    print(f"Parsing {filename}...")
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active  # Assume first sheet

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)

    if not header:
        print("❌ Empty file")
        return []

    # Detect columns
    sku_col_idx = -1
    name_col_idx = -1
    qty_col_idx = -1

    print(f"Headers: {header}")

    for idx, val in enumerate(header):
        if not val:
            continue
        val = str(val).lower().strip()
        if any(x in val for x in ["артикул", "sku", "код", "арт"]):
            sku_col_idx = idx
        elif any(x in val for x in ["название", "наименование", "name", "товар"]):
            name_col_idx = idx
        elif any(x in val for x in ["количество", "кол-во", "qty", "шт"]):
            qty_col_idx = idx

    print(f"Columns found: SKU={sku_col_idx}, Name={name_col_idx}, Qty={qty_col_idx}")

    if sku_col_idx == -1 and name_col_idx == -1:
        # Fallback to 0 and 1 if header detection failed (and maybe header is missing?)
        # For simplicity in prototype, we require headers or fallback logic similar to original
        # Original fallback: sku=0, name=None, qty=1
        if len(header) >= 1:
            sku_col_idx = 0
        if len(header) >= 2:
            qty_col_idx = 1
        print("⚠️ Used fallback column mapping")

    items = []
    for row in rows:
        if not row:
            continue

        sku = (
            str(row[sku_col_idx]).strip()
            if sku_col_idx != -1
            and sku_col_idx < len(row)
            and row[sku_col_idx] is not None
            else ""
        )
        name = (
            str(row[name_col_idx]).strip()
            if name_col_idx != -1
            and name_col_idx < len(row)
            and row[name_col_idx] is not None
            else ""
        )
        qty_val = (
            row[qty_col_idx] if qty_col_idx != -1 and qty_col_idx < len(row) else 1
        )

        try:
            qty = float(qty_val) if qty_val is not None else 1.0
        except (ValueError, TypeError):
            qty = 1.0

        if sku or name:
            items.append({"sku": sku, "name": name, "qty": qty})

    return items


def test_prototype():
    # 1. Create sample
    excel_file = create_sample_excel()

    # 2. Parse
    items = parse_excel_file(excel_file, "order.xlsx")

    # 3. Verify
    print(f"\nParsed {len(items)} items:")
    for item in items:
        print(item)

    assert len(items) == 2
    assert items[0]["sku"] == "101105032R"
    assert items[0]["qty"] == 10.0
    print("\n✅ Verification passed!")


if __name__ == "__main__":
    test_prototype()
