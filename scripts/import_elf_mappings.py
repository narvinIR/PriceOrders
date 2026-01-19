#!/usr/bin/env python3
"""Импорт маппингов клиента Эльф из Excel в БД

Использование:
    PYTHONPATH=. python scripts/import_elf_mappings.py
"""

import sys
sys.path.insert(0, '.')

from uuid import UUID
from openpyxl import load_workbook
from backend.models.database import get_supabase_client
from backend.services.matching import MatchingService


def main():
    db = get_supabase_client()

    # 1. Найти или создать клиента "Эльф"
    print("Поиск клиента Эльф...")
    client_resp = db.table('clients').select('id').eq('code', 'ELF').execute()
    
    if client_resp.data:
        client_id = UUID(client_resp.data[0]['id'])
        print(f"Клиент найден: {client_id}")
    else:
        print("Создание клиента Эльф...")
        client_resp = db.table('clients').insert({
            'name': 'Эльф',
            'code': 'ELF'
        }).execute()
        client_id = UUID(client_resp.data[0]['id'])
        print(f"Клиент создан: {client_id}")

    # 2. Загрузить все товары для поиска
    print("Загрузка каталога товаров...")
    products_resp = db.table('products').select('id, name').execute()
    products_by_name = {p['name']: p['id'] for p in products_resp.data}
    print(f"Загружено товаров: {len(products_by_name)}")

    # 3. Читать Excel и создавать маппинги
    print("Чтение Excel файла...")
    wb = load_workbook('ЭЛЬФ КАН январь 2026.xlsx')
    ws = wb['TDSheet']

    matcher = MatchingService()
    imported = 0
    skipped = 0
    errors = []

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:  # заголовок
            continue

        _, elf_name, article, jakko_name, _ = row[:5]

        # Пропускаем строки без артикула или названия Jakko
        if not article or not jakko_name:
            skipped += 1
            continue

        jakko_name = jakko_name.strip()

        # Найти product_id по точному названию
        product_id = products_by_name.get(jakko_name)

        # Если не найдено - попробовать частичное совпадение
        if not product_id:
            for name, pid in products_by_name.items():
                # Проверяем вхождение в обе стороны (без пробелов на концах)
                if jakko_name in name or name in jakko_name:
                    product_id = pid
                    break

        if not product_id:
            errors.append(f"Строка {i}: не найден товар '{jakko_name}'")
            continue

        # Сохраняем маппинг
        matcher.save_mapping(
            client_id=client_id,
            client_sku=article.strip(),
            product_id=UUID(product_id),
            confidence=100.0,
            match_type='excel_import',
            verified=True
        )
        imported += 1

    # Результаты
    print()
    print("=" * 50)
    print(f"Импортировано маппингов: {imported}")
    print(f"Пропущено (нет данных): {skipped}")

    if errors:
        print(f"Ошибки ({len(errors)}):")
        for e in errors[:20]:  # Показываем первые 20
            print(f"  - {e}")
        if len(errors) > 20:
            print(f"  ... и ещё {len(errors) - 20} ошибок")
    else:
        print("Ошибок нет!")

    print()
    print(f"Client ID: {client_id}")
    print("Проверить маппинги: GET /clients/{client_id}/mappings?verified_only=true")


if __name__ == '__main__':
    main()
