import csv
import re
from io import BytesIO, StringIO
from typing import BinaryIO

import openpyxl
from openpyxl.utils import get_column_letter

from backend.models.schemas import OrderItemBase
from backend.services.ocr_service import get_ocr_service  # lazy import if needed


def extract_pack_qty(name: str) -> int:
    """Извлечь количество в упаковке из названия товара"""
    if not name:
        return 1
    patterns = [
        r"\(уп\.?\s*(\d+)\s*шт\.?\)",  # (уп 20 шт), (уп.20шт)
        r"\((\d+)\s*шт\)",  # (20 шт)
        r"(\d+)\s*шт/кор",  # 100 шт/кор
        r"уп\.?\s*(\d+)\s*шт",  # уп 20 шт
        r"\((\d+)\s*шт/кор\)",  # (25 шт/кор)
    ]
    for pattern in patterns:
        match = re.search(pattern, name, re.IGNORECASE)
        if match:
            try:
                return int(match.group(1))
            except (ValueError, TypeError):
                pass
    return 1


def extract_thread_type(name: str) -> str | None:
    """Извлечь тип резьбы из названия товара"""
    if not name:
        return None
    name_lower = name.lower()
    if "вн.рез" in name_lower or "вн рез" in name_lower or "внутр" in name_lower:
        return "внутренняя"
    if "нар.рез" in name_lower or "нар рез" in name_lower or "наруж" in name_lower:
        return "наружная"
    return None


class ExcelService:
    """Парсинг Excel файлов без использования Pandas (для экономии памяти)"""

    SKU_COLUMNS = ["артикул", "sku", "код", "article", "code", "номер", "арт", "арт."]
    NAME_COLUMNS = ["название", "наименование", "name", "товар", "product", "описание"]
    QTY_COLUMNS = ["количество", "qty", "quantity", "кол-во", "кол", "шт", "count"]

    @classmethod
    def _read_csv(cls, file: BinaryIO) -> list[dict]:
        """Чтение CSV файла с автоопределением кодировки"""
        content = file.read()
        file.seek(0)

        encodings = ["utf-8", "cp1251", "latin-1", "utf-8-sig"]
        decoded_file = None

        for enc in encodings:
            try:
                text = content.decode(enc)
                decoded_file = StringIO(text)
                break
            except UnicodeDecodeError:
                continue

        if not decoded_file:
            raise ValueError("Не удалось определить кодировку CSV")

        # Detect delimiter
        try:
            sample = decoded_file.read(1024)
            decoded_file.seek(0)
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        except csv.Error:
            dialect = csv.excel

        reader = csv.DictReader(decoded_file, dialect=dialect)
        # Normalize headers
        if reader.fieldnames:
            reader.fieldnames = [str(f).lower().strip() for f in reader.fieldnames]

        return list(reader)

    @classmethod
    def _read_excel(cls, file: BinaryIO) -> list[dict]:
        """Чтение Excel файла через openpyxl"""
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active

        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return []

        if not header:
            return []

        # Normalize headers
        headers = [
            str(h).lower().strip() if h is not None else f"col_{i}"
            for i, h in enumerate(header)
        ]

        data = []
        for row in rows:
            if not any(row):
                continue
            item = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    item[headers[i]] = val
            data.append(item)

        wb.close()
        return data

    @classmethod
    def parse_order_file(cls, file: BinaryIO, filename: str) -> list[OrderItemBase]:
        """Парсинг файла заказа"""
        if filename.lower().endswith(".csv"):
            data = cls._read_csv(file)
        else:
            data = cls._read_excel(file)

        if not data:
            return []

        # Get columns from first row matching known headers
        # Note: data is list of dicts with lowered keys
        first_row = data[0] if data else {}
        columns = list(first_row.keys())

        sku_col = cls._find_column(columns, cls.SKU_COLUMNS)
        name_col = cls._find_column(columns, cls.NAME_COLUMNS)
        qty_col = cls._find_column(columns, cls.QTY_COLUMNS)

        if not sku_col and not name_col:
            # Fallback: use first column as SKU if unnamed
            if columns:
                sku_col = columns[0]

        items = []
        for row in data:
            sku_val = row.get(sku_col) if sku_col else None
            name_val = row.get(name_col) if name_col else None
            qty_val = row.get(qty_col) if qty_col else 1

            sku = str(sku_val).strip() if sku_val is not None else ""
            name = str(name_val).strip() if name_val is not None else ""

            try:
                qty = float(qty_val) if qty_val is not None else 1.0
            except (ValueError, TypeError):
                qty = 1.0

            if not sku and not name:
                continue

            items.append(
                OrderItemBase(
                    client_sku=sku or name[:50], client_name=name, quantity=qty
                )
            )

        return items

    @classmethod
    def _find_column(cls, columns: list, candidates: list) -> str | None:
        for col in columns:
            col_lower = col.lower()
            for cand in candidates:
                if cand in col_lower:
                    return col
        return None

    @classmethod
    def export_order(
        cls, order_data: list[dict], include_mapping: bool = True
    ) -> bytes:
        """Экспорт заказа в Excel (без pandas)"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Заказ"

        headers = ["Артикул клиента", "Название клиента", "Количество"]
        if include_mapping:
            headers.extend(
                [
                    "Артикул поставщика",
                    "Название поставщика",
                    "Упаковка",
                    "Совпадение %",
                    "Тип маппинга",
                    "Требует проверки",
                ]
            )

        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        for item in order_data:
            qty = item.get("quantity", 1)
            row = [item.get("client_sku", ""), item.get("client_name", ""), qty]

            if include_mapping:
                match = item.get("match", {})
                row.extend(
                    [
                        match.get("product_sku", ""),
                        match.get("product_name", ""),
                        match.get("pack_qty", 1),
                        match.get("confidence", 0),
                        match.get("match_type", ""),
                        "Да" if match.get("needs_review", True) else "Нет",
                    ]
                )
            ws.append(row)

        # Auto-width
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 0), len(str(cell.value)))
                    )
        for col, value in dims.items():
            ws.column_dimensions[col].width = min(value + 2, 50)  # Cap width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @classmethod
    def parse_jakko_catalog(cls, file: BinaryIO) -> list[dict]:
        """Парсинг каталога Jakko (через openpyxl)"""
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        products = []

        categories = {
            "1": "Трубы ПЭ для воды",
            "2": "Трубы PERT",
            "3": "Трубы ППР",
            "4": "Фитинги ППР",
            "5": "Фитинги ППР с резьбой",
            "6": "Запорная арматура ППР",
            "7": "Трубы ПП малошумные",
            "8": "Трубы канализационные ПП",
            "9": "Трубы наружной канализации",
            "10": "Трубы рифлёные",
            "11": "Герметики и инструмент",
            "12": "Прочее",
        }

        for sheet_name in wb.sheetnames:
            if sheet_name in ["Содержание", "заказ"]:
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if len(rows) < 5:
                continue

            # Header search
            header_row = rows[4]  # 0-indexed, row 5
            sku_idx = -1
            name_idx = -1

            for idx, val in enumerate(header_row):
                val_str = str(val).upper().strip() if val else ""
                if "АРТИКУЛ" in val_str:
                    sku_idx = idx
                elif "НОМЕНКЛАТУРА" in val_str:
                    name_idx = idx

            if sku_idx == -1 or name_idx == -1:
                continue

            for i in range(5, len(rows)):
                row = rows[i]
                if sku_idx >= len(row) or name_idx >= len(row):
                    continue

                sku = str(row[sku_idx]).strip() if row[sku_idx] else ""
                name = str(row[name_idx]).strip() if row[name_idx] else ""

                if not sku or sku == "None" or sku == "АРТИКУЛ":
                    continue
                name = name.replace("\xa0", " ").strip()
                if not name:
                    continue

                products.append(
                    {
                        "sku": sku,
                        "name": name,
                        "category": categories.get(sheet_name, sheet_name),
                        "pack_qty": extract_pack_qty(name),
                        "attributes": {"thread_type": extract_thread_type(name)},
                    }
                )

        wb.close()
        return products
